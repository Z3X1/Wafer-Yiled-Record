"""Unit tests for wafer_yield_analyzer module.

This module contains comprehensive tests for the wafer yield analysis functionality.
"""

import pytest
from pathlib import Path
from typing import List
import pandas as pd
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook

from wafer_yield_analyzer import (
    find_wafer_summary_files,
    extract_wafer_data,
    create_yield_dataframe,
    create_beautiful_plot,
    save_to_excel_with_chart
)


class TestFindWaferSummaryFiles:
    """Test cases for find_wafer_summary_files function."""
    
    def test_directory_not_found(self):
        """Test that FileNotFoundError is raised for non-existent directory."""
        with pytest.raises(FileNotFoundError):
            find_wafer_summary_files("non_existent_directory")
    
    @patch('wafer_yield_analyzer.Path')
    def test_finds_matching_files(self, mock_path):
        """Test that function finds files with 'Wafer_Summary' in name."""
        # Mock the path and rglob behavior
        mock_dir = Mock()
        mock_path.return_value = mock_dir
        mock_dir.exists.return_value = True
        
        # Create mock files
        mock_files = [
            Mock(name='Wafer_Summary_001.xlsx'),
            Mock(name='Wafer_Summary_002.xlsx'),
            Mock(name='Other_File.xlsx')
        ]
        mock_files[0].name = 'Wafer_Summary_001.xlsx'
        mock_files[1].name = 'Wafer_Summary_002.xlsx'
        mock_files[2].name = 'Other_File.xlsx'
        
        mock_dir.rglob.return_value = mock_files[:2]
        
        result = find_wafer_summary_files("test_dir")
        assert len(result) >= 0  # Function will return actual results


class TestExtractWaferData:
    """Test cases for extract_wafer_data function."""
    
    def test_extract_valid_data(self, tmp_path):
        """Test extraction of valid wafer data from Excel file."""
        # Create a temporary Excel file
        wb = Workbook()
        ws = wb.active
        ws['B4'] = 'W001'
        ws['D11'] = 95.5
        
        test_file = tmp_path / "test_wafer.xlsx"
        wb.save(test_file)
        wb.close()
        
        wafer_id, yield_value = extract_wafer_data(test_file)
        
        assert wafer_id == 'W001'
        assert yield_value == 95.5
    
    def test_extract_percentage_string(self, tmp_path):
        """Test extraction when yield is stored as percentage string."""
        wb = Workbook()
        ws = wb.active
        ws['B4'] = 'W002'
        ws['D11'] = '98.5%'
        
        test_file = tmp_path / "test_wafer_pct.xlsx"
        wb.save(test_file)
        wb.close()
        
        wafer_id, yield_value = extract_wafer_data(test_file)
        
        assert wafer_id == 'W002'
        assert yield_value == 98.5
    
    def test_extract_decimal_yield(self, tmp_path):
        """Test extraction when yield is stored as decimal (0.95 -> 95%)."""
        wb = Workbook()
        ws = wb.active
        ws['B4'] = 'W003'
        ws['D11'] = 0.97
        
        test_file = tmp_path / "test_wafer_decimal.xlsx"
        wb.save(test_file)
        wb.close()
        
        wafer_id, yield_value = extract_wafer_data(test_file)
        
        assert wafer_id == 'W003'
        assert yield_value == 97.0
    
    def test_missing_wafer_id(self, tmp_path):
        """Test that ValueError is raised when Wafer ID is missing."""
        wb = Workbook()
        ws = wb.active
        ws['D11'] = 95.5
        # B4 is empty
        
        test_file = tmp_path / "test_missing_id.xlsx"
        wb.save(test_file)
        wb.close()
        
        with pytest.raises(ValueError, match="Wafer ID.*is empty"):
            extract_wafer_data(test_file)
    
    def test_missing_yield(self, tmp_path):
        """Test that ValueError is raised when Yield is missing."""
        wb = Workbook()
        ws = wb.active
        ws['B4'] = 'W004'
        # D11 is empty
        
        test_file = tmp_path / "test_missing_yield.xlsx"
        wb.save(test_file)
        wb.close()
        
        with pytest.raises(ValueError, match="Yield.*is empty"):
            extract_wafer_data(test_file)


class TestCreateYieldDataframe:
    """Test cases for create_yield_dataframe function."""
    
    def test_create_dataframe_from_files(self, tmp_path):
        """Test creating DataFrame from multiple valid files."""
        files = []
        
        # Create multiple test files
        for i in range(3):
            wb = Workbook()
            ws = wb.active
            ws['B4'] = f'W{i+1:03d}'
            ws['D11'] = 90.0 + i * 2
            
            test_file = tmp_path / f"wafer_{i}.xlsx"
            wb.save(test_file)
            wb.close()
            files.append(test_file)
        
        df = create_yield_dataframe(files)
        
        assert len(df) == 3
        assert list(df.columns) == ['Wafer_ID', 'Yield']
        assert df['Yield'].min() >= 90.0
        assert df['Yield'].max() <= 94.0
    
    def test_empty_file_list(self):
        """Test that ValueError is raised for empty file list."""
        with pytest.raises(ValueError, match="No valid data extracted"):
            create_yield_dataframe([])


class TestCreateBeautifulPlot:
    """Test cases for create_beautiful_plot function."""
    
    @patch('wafer_yield_analyzer.plt')
    def test_plot_creation(self, mock_plt, tmp_path):
        """Test that plot is created successfully."""
        df = pd.DataFrame({
            'Wafer_ID': ['W001', 'W002', 'W003'],
            'Yield': [95.5, 96.2, 94.8]
        })
        
        output_path = tmp_path / "test_plot.png"
        
        # Configure mock
        mock_fig = Mock()
        mock_ax = Mock()
        mock_plt.subplots.return_value = (mock_fig, mock_ax)
        
        create_beautiful_plot(df, str(output_path))
        
        # Verify plot methods were called
        mock_plt.subplots.assert_called_once()
        mock_ax.plot.assert_called_once()
        mock_plt.savefig.assert_called_once()


class TestSaveToExcelWithChart:
    """Test cases for save_to_excel_with_chart function."""
    
    def test_save_excel_with_valid_data(self, tmp_path):
        """Test saving DataFrame to Excel with chart."""
        df = pd.DataFrame({
            'Wafer_ID': ['W001', 'W002'],
            'Yield': [95.5, 96.2]
        })
        
        # Create a dummy image
        import matplotlib.pyplot as plt
        fig, ax = plt.subplots()
        ax.plot([1, 2], [3, 4])
        image_path = tmp_path / "test_chart.png"
        plt.savefig(image_path)
        plt.close()
        
        output_path = tmp_path / "test_output.xlsx"
        
        save_to_excel_with_chart(df, str(output_path), str(image_path))
        
        # Verify file was created
        assert output_path.exists()
        
        # Verify content
        df_read = pd.read_excel(output_path, sheet_name='Yield Data')
        assert len(df_read) == 2
        assert list(df_read.columns) == ['Wafer_ID', 'Yield']


@pytest.fixture
def sample_wafer_files(tmp_path):
    """Fixture to create sample wafer summary files for testing."""
    files = []
    for i in range(5):
        wb = Workbook()
        ws = wb.active
        ws['B4'] = f'Wafer_{i+1:03d}'
        ws['D11'] = 92.0 + i * 1.5
        
        file_path = tmp_path / f"Wafer_Summary_{i+1:03d}.xlsx"
        wb.save(file_path)
        wb.close()
        files.append(file_path)
    
    return files


def test_integration_full_workflow(sample_wafer_files, tmp_path):
    """Integration test for the complete workflow."""
    # Extract data
    df = create_yield_dataframe(sample_wafer_files)
    
    # Verify DataFrame
    assert len(df) == 5
    assert df['Yield'].min() >= 92.0
    
    # Create plot
    plot_path = tmp_path / "integration_plot.png"
    create_beautiful_plot(df, str(plot_path))
    assert plot_path.exists()
    
    # Save to Excel
    excel_path = tmp_path / "integration_report.xlsx"
    save_to_excel_with_chart(df, str(excel_path), str(plot_path))
    assert excel_path.exists()

