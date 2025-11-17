"""Wafer Yield Analyzer.

This module reads Excel files containing wafer summary data and generates
a yield analysis report with visualization.
"""

from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl import load_workbook
import seaborn as sns
import matplotlib.pyplot as plt
from pathlib import Path
from typing import List, Tuple
import logging

import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend for better compatibility

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Set style for better visualization
try:
    plt.style.use('seaborn-v0_8-darkgrid')
except OSError:
    # Fallback to default style if seaborn style is not available
    plt.style.use('default')
sns.set_palette("husl")


def find_wafer_summary_files(directory: str) -> List[Path]:
    """Find all Excel files containing 'Wafer_Summary' in the filename.

    Args:
        directory: The directory path to search for Excel files.

    Returns:
        A list of Path objects for matching Excel files.

    Raises:
        FileNotFoundError: If the directory does not exist.
    """
    dir_path = Path(directory)

    if not dir_path.exists():
        raise FileNotFoundError(f"Directory not found: {directory}")

    # Search for Excel files with 'Wafer_Summary' in filename
    excel_patterns = ['*.xlsx', '*.xls']
    matching_files = []

    for pattern in excel_patterns:
        for file_path in dir_path.rglob(pattern):
            # Skip temporary Excel files (starting with ~$)
            if file_path.name.startswith('~$'):
                continue
            if 'Wafer_Summary' in file_path.name:
                matching_files.append(file_path)

    logger.info(f"Found {len(matching_files)} Wafer_Summary files")
    return matching_files


def extract_wafer_data(file_path: Path) -> Tuple[str, float]:
    """Extract Wafer ID and Yield from an Excel file.

    Args:
        file_path: Path to the Excel file.

    Returns:
        A tuple containing (Wafer ID, Yield percentage).

    Raises:
        ValueError: If required cells are empty or invalid.
    """
    try:
        # Read specific cells using openpyxl for precise cell access
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Extract Wafer ID from B4
        wafer_id = ws['B4'].value

        # Extract Yield from D11 (assuming it's a percentage)
        yield_value = ws['D11'].value

        wb.close()

        # Validate data
        if wafer_id is None:
            raise ValueError(f"Wafer ID (B4) is empty in {file_path.name}")

        if yield_value is None:
            raise ValueError(f"Yield (D11) is empty in {file_path.name}")

        # Convert yield to float if it's a string with '%'
        if isinstance(yield_value, str):
            yield_value = float(yield_value.strip('%'))
        else:
            yield_value = float(yield_value)
            # If the value is like 0.95, convert to percentage (95)
            if yield_value <= 1.0:
                yield_value *= 100

        logger.info(
            f"Extracted data from {file_path.name}: ID={wafer_id}, Yield={yield_value}%")
        return str(wafer_id), yield_value

    except Exception as e:
        logger.error(f"Error processing {file_path.name}: {e}")
        raise


def create_yield_dataframe(files: List[Path]) -> pd.DataFrame:
    """Create a DataFrame with Wafer ID and Yield data.

    Args:
        files: List of Excel file paths to process.

    Returns:
        A pandas DataFrame with columns ['Wafer_ID', 'Yield'].
    """
    data = []

    for file_path in files:
        try:
            wafer_id, yield_value = extract_wafer_data(file_path)
            data.append({'Wafer_ID': wafer_id, 'Yield': yield_value})
        except Exception as e:
            logger.warning(f"Skipping {file_path.name} due to error: {e}")
            continue

    if not data:
        raise ValueError("No valid data extracted from any files")

    df = pd.DataFrame(data)
    # Sort by Wafer ID for better visualization
    df = df.sort_values('Wafer_ID').reset_index(drop=True)

    return df


def create_beautiful_plot(df: pd.DataFrame, output_image: str) -> None:
    """Create a beautiful line chart for wafer yield data.

    Args:
        df: DataFrame with Wafer_ID and Yield columns.
        output_image: Path to save the plot image.
    """
    try:
        fig, ax = plt.subplots(figsize=(14, 8))
    except Exception as e:
        logger.error(f"Error creating figure: {e}")
        # Use simple style as fallback
        plt.style.use('default')
        fig, ax = plt.subplots(figsize=(14, 8))

    # Plot line with markers
    ax.plot(df['Wafer_ID'], df['Yield'],
            marker='o',
            markersize=10,
            linewidth=2.5,
            color='#2E86AB',
            markerfacecolor='#A23B72',
            markeredgecolor='white',
            markeredgewidth=2,
            label='Wafer Yield')

    # Add value labels on each point
    for idx, row in df.iterrows():
        ax.annotate(f'{row["Yield"]:.2f}%',
                    (row['Wafer_ID'], row['Yield']),
                    textcoords="offset points",
                    xytext=(0, 10),
                    ha='center',
                    fontsize=9,
                    fontweight='bold',
                    color='#333333')

    # Styling
    ax.set_xlabel('Wafer ID', fontsize=14, fontweight='bold', color='#333333')
    ax.set_ylabel('Yield (%)', fontsize=14, fontweight='bold', color='#333333')
    ax.set_title('Wafer Yield Analysis',
                 fontsize=18,
                 fontweight='bold',
                 color='#2E86AB',
                 pad=20)

    # Grid styling
    ax.grid(True, linestyle='--', alpha=0.6, linewidth=0.8)
    ax.set_axisbelow(True)

    # Set y-axis limits with some padding
    y_min = max(0, df['Yield'].min() - 5)
    y_max = min(100, df['Yield'].max() + 5)
    ax.set_ylim(y_min, y_max)

    # Rotate x-axis labels for better readability
    plt.xticks(rotation=45, ha='right', fontsize=10)
    plt.yticks(fontsize=10)

    # Add legend
    ax.legend(loc='best', fontsize=11, frameon=True, shadow=True)

    # Add a subtle background color
    ax.set_facecolor('#F8F9FA')
    fig.patch.set_facecolor('white')

    # Tight layout to prevent label cutoff
    plt.tight_layout()

    # Save the figure
    try:
        plt.savefig(output_image, dpi=300,
                    bbox_inches='tight', facecolor='white')
        logger.info(f"✓ Plot saved successfully: {output_image}")

        # Verify the file was created
        if Path(output_image).exists():
            file_size = Path(output_image).stat().st_size
            logger.info(f"Image file size: {file_size / 1024:.2f} KB")
        else:
            logger.error(f"Image file was not created: {output_image}")
            raise FileNotFoundError(
                f"Failed to create image file: {output_image}")
    except Exception as e:
        logger.error(f"Error saving plot: {e}", exc_info=True)
        raise
    finally:
        plt.close()


def save_to_excel_with_chart(df: pd.DataFrame, output_path: str, image_path: str) -> None:
    """Save the DataFrame to Excel with an embedded chart.

    Args:
        df: DataFrame with Wafer_ID and Yield columns.
        output_path: Path to save the Excel file.
        image_path: Path to the chart image to embed.
    """
    try:
        # Verify image file exists
        if not Path(image_path).exists():
            logger.error(f"Image file not found: {image_path}")
            raise FileNotFoundError(f"Image file not found: {image_path}")

        logger.info(f"Creating Excel file: {output_path}")
        logger.info(f"Embedding chart from: {image_path}")

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write data to Excel
            df.to_excel(writer, sheet_name='Yield Data', index=False)
            logger.info("Data written to 'Yield Data' sheet")

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Yield Data']

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            logger.info("Column widths adjusted")

            # Create a new sheet for the chart
            chart_sheet = workbook.create_sheet('Yield Chart')
            logger.info("Created 'Yield Chart' sheet")

            # Insert the matplotlib image
            from openpyxl.drawing.image import Image
            img = Image(image_path)
            img.width = 1000
            img.height = 571
            chart_sheet.add_image(img, 'A1')
            logger.info("Chart image embedded successfully")

        logger.info(f"✓ Excel file saved successfully: {output_path}")

    except Exception as e:
        logger.error(f"Error saving Excel file: {e}", exc_info=True)
        raise


def main() -> None:
    """Main function to orchestrate the wafer yield analysis."""
    try:
        # Configuration
        source_directory = r"C:\Users\andrel52"
        output_excel = "wafer_yield_report.xlsx"
        output_image = "wafer_yield_chart.png"

        logger.info("Starting Wafer Yield Analysis...")

        # Find all Wafer_Summary files
        wafer_files = find_wafer_summary_files(source_directory)

        if not wafer_files:
            logger.error(f"No Wafer_Summary files found in {source_directory}")
            return

        # Extract data and create DataFrame
        logger.info("Extracting wafer data...")
        df = create_yield_dataframe(wafer_files)

        logger.info(f"Successfully processed {len(df)} wafers")
        logger.info(f"\nSummary Statistics:\n{df.describe()}")

        # Create visualization
        logger.info("Creating visualization...")
        create_beautiful_plot(df, output_image)

        # Save to Excel
        logger.info("Saving to Excel...")
        save_to_excel_with_chart(df, output_excel, output_image)

        logger.info("=" * 60)
        logger.info("✓ Analysis completed successfully!")
        logger.info(f"✓ Excel report: {output_excel}")
        logger.info(f"✓ Chart image: {output_image}")
        logger.info("=" * 60)

    except Exception as e:
        logger.error(f"Analysis failed: {e}", exc_info=True)
        raise


if __name__ == "__main__":
    main()
