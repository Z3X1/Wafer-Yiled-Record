"""Advanced Wafer Yield Analyzer with Configuration Support.

This module provides an enhanced version of the wafer yield analyzer
with YAML configuration support and additional features.
"""

from pathlib import Path
from typing import List, Tuple, Dict, Any
import logging

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import yaml

# Configure logging (will be updated from config)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def load_config(config_path: str = "config.yaml") -> Dict[str, Any]:
    """Load configuration from YAML file.
    
    Args:
        config_path: Path to the configuration file.
        
    Returns:
        Dictionary containing configuration settings.
        
    Raises:
        FileNotFoundError: If config file doesn't exist.
        yaml.YAMLError: If config file is invalid.
    """
    config_file = Path(config_path)
    
    if not config_file.exists():
        logger.warning(f"Config file not found: {config_path}, using defaults")
        return get_default_config()
    
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        logger.info(f"Loaded configuration from {config_path}")
        return config
    except yaml.YAMLError as e:
        logger.error(f"Error parsing config file: {e}")
        raise


def get_default_config() -> Dict[str, Any]:
    """Get default configuration settings.
    
    Returns:
        Dictionary with default configuration values.
    """
    return {
        'input': {
            'source_directory': r'C:\Users\andrel52',
            'wafer_id_cell': 'B4',
            'yield_cell': 'D11',
            'file_pattern': 'Wafer_Summary',
            'extensions': ['*.xlsx', '*.xls']
        },
        'output': {
            'excel_filename': 'wafer_yield_report.xlsx',
            'chart_filename': 'wafer_yield_chart.png',
            'output_directory': '.',
            'data_sheet_name': 'Yield Data',
            'chart_sheet_name': 'Yield Chart'
        },
        'visualization': {
            'figure_size': {'width': 14, 'height': 8},
            'dpi': 300,
            'line': {'color': '#2E86AB', 'width': 2.5},
            'marker': {
                'shape': 'o',
                'size': 10,
                'facecolor': '#A23B72',
                'edgecolor': 'white',
                'edgewidth': 2
            },
            'chart': {
                'title': 'Wafer Yield Analysis',
                'title_fontsize': 18,
                'xlabel': 'Wafer ID',
                'ylabel': 'Yield (%)',
                'label_fontsize': 14,
                'grid_alpha': 0.6,
                'background_color': '#F8F9FA'
            },
            'y_axis_padding': 5
        },
        'processing': {
            'sort_by_id': True,
            'skip_on_error': True,
            'validate_yield_range': True,
            'min_yield': 0.0,
            'max_yield': 100.0
        }
    }


def find_wafer_summary_files(config: Dict[str, Any]) -> List[Path]:
    """Find all Excel files matching the pattern.
    
    Args:
        config: Configuration dictionary.
        
    Returns:
        List of Path objects for matching Excel files.
        
    Raises:
        FileNotFoundError: If the directory does not exist.
    """
    input_config = config['input']
    dir_path = Path(input_config['source_directory'])
    
    if not dir_path.exists():
        raise FileNotFoundError(f"Directory not found: {dir_path}")
    
    pattern = input_config['file_pattern']
    extensions = input_config['extensions']
    matching_files = []
    
    for ext_pattern in extensions:
        for file_path in dir_path.rglob(ext_pattern):
            if pattern in file_path.name:
                matching_files.append(file_path)
    
    logger.info(f"Found {len(matching_files)} files matching '{pattern}'")
    return matching_files


def extract_wafer_data(
    file_path: Path,
    config: Dict[str, Any]
) -> Tuple[str, float]:
    """Extract Wafer ID and Yield from an Excel file.
    
    Args:
        file_path: Path to the Excel file.
        config: Configuration dictionary.
        
    Returns:
        Tuple containing (Wafer ID, Yield percentage).
        
    Raises:
        ValueError: If required cells are empty or invalid.
    """
    input_config = config['input']
    processing_config = config['processing']
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Extract data using configured cell locations
        wafer_id = ws[input_config['wafer_id_cell']].value
        yield_value = ws[input_config['yield_cell']].value
        
        wb.close()
        
        # Validate data
        if wafer_id is None:
            raise ValueError(
                f"Wafer ID ({input_config['wafer_id_cell']}) is empty "
                f"in {file_path.name}"
            )
        
        if yield_value is None:
            raise ValueError(
                f"Yield ({input_config['yield_cell']}) is empty "
                f"in {file_path.name}"
            )
        
        # Convert yield to float
        if isinstance(yield_value, str):
            yield_value = float(yield_value.strip('%'))
        else:
            yield_value = float(yield_value)
            if yield_value <= 1.0:
                yield_value *= 100
        
        # Validate yield range if configured
        if processing_config['validate_yield_range']:
            min_yield = processing_config['min_yield']
            max_yield = processing_config['max_yield']
            if not (min_yield <= yield_value <= max_yield):
                logger.warning(
                    f"Yield {yield_value}% out of range [{min_yield}, {max_yield}] "
                    f"in {file_path.name}"
                )
        
        logger.info(
            f"Extracted: {file_path.name} -> ID={wafer_id}, Yield={yield_value:.2f}%"
        )
        return str(wafer_id), yield_value
        
    except Exception as e:
        logger.error(f"Error processing {file_path.name}: {e}")
        if not processing_config['skip_on_error']:
            raise
        raise


def create_yield_dataframe(
    files: List[Path],
    config: Dict[str, Any]
) -> pd.DataFrame:
    """Create a DataFrame with Wafer ID and Yield data.
    
    Args:
        files: List of Excel file paths to process.
        config: Configuration dictionary.
        
    Returns:
        A pandas DataFrame with columns ['Wafer_ID', 'Yield'].
    """
    processing_config = config['processing']
    data = []
    
    for file_path in files:
        try:
            wafer_id, yield_value = extract_wafer_data(file_path, config)
            data.append({'Wafer_ID': wafer_id, 'Yield': yield_value})
        except Exception as e:
            if processing_config['skip_on_error']:
                logger.warning(f"Skipping {file_path.name}: {e}")
                continue
            else:
                raise
    
    if not data:
        raise ValueError("No valid data extracted from any files")
    
    df = pd.DataFrame(data)
    
    # Sort by Wafer ID if configured
    if processing_config['sort_by_id']:
        df = df.sort_values('Wafer_ID').reset_index(drop=True)
    
    return df


def create_beautiful_plot(
    df: pd.DataFrame,
    config: Dict[str, Any]
) -> str:
    """Create a beautiful line chart for wafer yield data.
    
    Args:
        df: DataFrame with Wafer_ID and Yield columns.
        config: Configuration dictionary.
        
    Returns:
        Path to the saved plot image.
    """
    viz_config = config['visualization']
    output_config = config['output']
    
    # Create figure
    fig_size = viz_config['figure_size']
    fig, ax = plt.subplots(figsize=(fig_size['width'], fig_size['height']))
    
    # Plot configuration
    line_config = viz_config['line']
    marker_config = viz_config['marker']
    
    # Plot line with markers
    ax.plot(
        df['Wafer_ID'],
        df['Yield'],
        marker=marker_config['shape'],
        markersize=marker_config['size'],
        linewidth=line_config['width'],
        color=line_config['color'],
        markerfacecolor=marker_config['facecolor'],
        markeredgecolor=marker_config['edgecolor'],
        markeredgewidth=marker_config['edgewidth'],
        label='Wafer Yield'
    )
    
    # Add value labels
    for idx, row in df.iterrows():
        ax.annotate(
            f'{row["Yield"]:.2f}%',
            (row['Wafer_ID'], row['Yield']),
            textcoords="offset points",
            xytext=(0, 10),
            ha='center',
            fontsize=9,
            fontweight='bold',
            color='#333333'
        )
    
    # Chart styling
    chart_config = viz_config['chart']
    ax.set_xlabel(
        chart_config['xlabel'],
        fontsize=chart_config['label_fontsize'],
        fontweight='bold',
        color='#333333'
    )
    ax.set_ylabel(
        chart_config['ylabel'],
        fontsize=chart_config['label_fontsize'],
        fontweight='bold',
        color='#333333'
    )
    ax.set_title(
        chart_config['title'],
        fontsize=chart_config['title_fontsize'],
        fontweight='bold',
        color=line_config['color'],
        pad=20
    )
    
    # Grid styling
    ax.grid(True, linestyle='--', alpha=chart_config['grid_alpha'], linewidth=0.8)
    ax.set_axisbelow(True)
    
    # Y-axis limits
    padding = viz_config['y_axis_padding']
    y_min = max(0, df['Yield'].min() - padding)
    y_max = min(100, df['Yield'].max() + padding)
    ax.set_ylim(y_min, y_max)
    
    # Styling
    plt.xticks(rotation=45, ha='right', fontsize=10)
    plt.yticks(fontsize=10)
    ax.legend(loc='best', fontsize=11, frameon=True, shadow=True)
    ax.set_facecolor(chart_config['background_color'])
    fig.patch.set_facecolor('white')
    
    plt.tight_layout()
    
    # Save figure
    output_dir = Path(output_config['output_directory'])
    output_dir.mkdir(parents=True, exist_ok=True)
    
    output_path = output_dir / output_config['chart_filename']
    plt.savefig(
        output_path,
        dpi=viz_config['dpi'],
        bbox_inches='tight',
        facecolor='white'
    )
    logger.info(f"Chart saved to {output_path}")
    
    plt.close()
    return str(output_path)


def save_to_excel_with_chart(
    df: pd.DataFrame,
    image_path: str,
    config: Dict[str, Any]
) -> str:
    """Save the DataFrame to Excel with an embedded chart.
    
    Args:
        df: DataFrame with Wafer_ID and Yield columns.
        image_path: Path to the chart image.
        config: Configuration dictionary.
        
    Returns:
        Path to the saved Excel file.
    """
    output_config = config['output']
    
    # Prepare output path
    output_dir = Path(output_config['output_directory'])
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_config['excel_filename']
    
    # Create Excel writer
    with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
        # Write data
        df.to_excel(
            writer,
            sheet_name=output_config['data_sheet_name'],
            index=False
        )
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[output_config['data_sheet_name']]
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create chart sheet
        chart_sheet = workbook.create_sheet(output_config['chart_sheet_name'])
        
        # Insert image
        from openpyxl.drawing.image import Image
        img = Image(image_path)
        img.width = 1000
        img.height = 571
        chart_sheet.add_image(img, 'A1')
    
    logger.info(f"Excel report saved to {output_path}")
    return str(output_path)


def main(config_path: str = "config.yaml") -> None:
    """Main function to orchestrate the wafer yield analysis.
    
    Args:
        config_path: Path to the configuration file.
    """
    try:
        # Load configuration
        logger.info("=" * 60)
        logger.info("Starting Advanced Wafer Yield Analysis")
        logger.info("=" * 60)
        
        config = load_config(config_path)
        
        # Find files
        wafer_files = find_wafer_summary_files(config)
        
        if not wafer_files:
            logger.error("No matching files found")
            return
        
        # Extract and process data
        logger.info("Extracting wafer data...")
        df = create_yield_dataframe(wafer_files, config)
        
        # Display statistics
        logger.info(f"\n{'=' * 60}")
        logger.info(f"Processed {len(df)} wafers successfully")
        logger.info(f"Average Yield: {df['Yield'].mean():.2f}%")
        logger.info(f"Max Yield: {df['Yield'].max():.2f}%")
        logger.info(f"Min Yield: {df['Yield'].min():.2f}%")
        logger.info(f"Std Deviation: {df['Yield'].std():.2f}%")
        logger.info(f"{'=' * 60}\n")
        
        # Create visualization
        logger.info("Creating visualization...")
        image_path = create_beautiful_plot(df, config)
        
        # Save to Excel
        logger.info("Generating Excel report...")
        excel_path = save_to_excel_with_chart(df, image_path, config)
        
        # Summary
        logger.info("=" * 60)
        logger.info("✓ Analysis completed successfully!")
        logger.info(f"✓ Excel report: {excel_path}")
        logger.info(f"✓ Chart image: {image_path}")
        logger.info("=" * 60)
        
    except Exception as e:
        logger.error(f"Analysis failed: {e}", exc_info=True)
        raise


if __name__ == "__main__":
    main()

